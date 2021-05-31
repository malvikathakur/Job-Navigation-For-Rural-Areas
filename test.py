from tkinter import *
import pandas as pd
import numpy as np
from gtts import gTTS
import openpyxl
import pathlib
from openpyxl import Workbook
import os
from tkinter import messagebox
import joblib
from tkinter.ttk import Combobox
import sqlite3
import speech_recognition as sr
from twilio.rest import Client
import random
file=pathlib.Path("Queries.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Title"
    file.save("Queries.xlsx")
clf = joblib.load('model.pickle')

conn = sqlite3.connect('Form.db')
with conn:
   cursor = conn.cursor()

def main2():
   root1 = Tk()
   root1.maxsize(350, 600)
   root1.minsize(350, 600)
   root1.configure(bg='white')
   frame = Frame(root1, highlightbackground="black", highlightthickness=25)
   frame.pack(side=TOP, expand=True, fill=BOTH)

   def cls():
      root1.destroy()
      main_()
   #conn = sqlite3.connect('Form.db')
   #with conn:
      #cursor = conn.cursor()

   def open():
      cursor.execute("select Fullname, Age, Contact, Salary, Location from User LIMIT 1 ")
      e = cursor.fetchall()
      name = e[0][0]
      print(name)
      age_ = e[0][1]
      salary_ = e[0][3]
      loc1 = e[0][4]
      d1 = {'col1': [age_], 'col2': [salary_]}
      df1 = pd.DataFrame(data=d1)
      result1 = clf.predict(df1)
      # print((type(result1)))
      result1 = np.array_str(result1)
      result1 = result1[2:-2]
      z = str(result1)
      # print(result1)
      txt = "Predicted job for " + e[0][0] + " is " + z

      label_01 = Label(root1, text=txt, width=30, bg='black', fg='white', font=("Calibri 10 bold", 10))
      label_01.place(x=40, y=220)
      answer = messagebox.askyesno("Check Vacancy", name + ", do you want to check vacancy in " + z + " job?")

      if answer:  # to check vacancy

         cursor.execute("select * from Employer where Age=(?) and Location=(?)", (result1, loc1,))
         e1 = cursor.fetchall()
         if not e1:#shift the cycle
            messagebox.showerror("Vacancy","Sorry no Vacancy Found. \nTry again later!")
            cursor.execute("select Fullname, Age, Contact, Salary, Location from User LIMIT 1 ")
            e11 = cursor.fetchall()
            name_ = e11[0][0]
            age_ = e11[0][1]
            contact_ = e11[0][2]
            salary_ = e11[0][3]
            loc = e11[0][4]
            x1 = (name_, age_, contact_, salary_, loc)
            cursor.execute("UPDATE User SET rowid = ((select max(rowid) from User)+1) WHERE rowid = 1")
            # conn.commit()
            cursor.execute("UPDATE User SET rowid = rowid - 1")
            # conn.commit()
            cursor.execute("select * from User order by rowid")
            messagebox.showinfo("After Updation", cursor.fetchall())
            print("This is also working.")
         else:
            nm = e1[0][0]
            cntct = e1[0][1]
            jbtp = e1[0][2]
            ag = e1[0][3]
            salr = e1[0][4]
            ans1 = messagebox.askokcancel("Job Found", "CONFIRM ?""\nJob " + result1 + "\nSalary " + e1[0][4])
            if ans1:  # deleting from user and employer table
               messagebox.showinfo("Contact Details", "Name:  "+e1[0][0]+"\nContact:  "+e1[0][1]+"\nSalary:  "+e1[0][4]+"\nLocation:  "+e1[0][5])
               cursor.execute("DELETE FROM User WHERE rowid in (select rowid FROM User LIMIT 1)")
               #conn.commit()
               cursor.execute("select * from User")
               messagebox.showinfo("User Update", cursor.fetchall())
               cursor.execute(
                  "DELETE FROM Employer WHERE Fullname=(?) and Contact=(?) and Age=(?) and JOB=(?) and Salary=(?)",
                  (nm, cntct, jbtp, ag, salr,))
               #conn.commit()
               cursor.execute("select * from Employer")
               messagebox.showinfo("Employer Update", cursor.fetchall())
            else:  # put that user at the bottom row shift cycle
               cursor.execute("select Fullname, Age, Contact, Salary, Location from User LIMIT 1 ")
               e11 = cursor.fetchall()
               name_ = e11[0][0]
               age_ = e11[0][1]
               contact_ = e11[0][2]
               salary_ = e11[0][3]
               loc = e11[0][4]
               x1 = (name_, age_, contact_, salary_, loc)
               cursor.execute("UPDATE User SET rowid = ((select max(rowid) from User)+1) WHERE rowid = 1")
               # conn.commit()
               cursor.execute("UPDATE User SET rowid = rowid - 1")
               # conn.commit()
               cursor.execute("select * from User order by rowid")
               messagebox.showinfo("After Updation", cursor.fetchall())
               print("no")
      else:  # putting the top row to the bottom for fcfs
         cursor.execute("select Fullname, Age, Contact, Salary, Location from User LIMIT 1 ")
         e11 = cursor.fetchall()
         name_ = e11[0][0]
         age_ = e11[0][1]
         contact_ = e11[0][2]
         salary_ = e11[0][3]
         loc = e11[0][4]
         x1 = (name_, age_, contact_, salary_, loc)
         cursor.execute("UPDATE User SET rowid = ((select max(rowid) from User)+1) WHERE rowid = 1")
         #conn.commit()
         cursor.execute("UPDATE User SET rowid = rowid - 1")
         #conn.commit()
         cursor.execute("select * from User order by rowid")
         messagebox.showinfo("After Updation", cursor.fetchall())
         print("no")

   Button(root1, text='START', command=open, width=10, bg='black', fg='white').place(x=130, y=180)

   bottomframe = Frame(frame, highlightbackground="black", highlightthickness=5)
   bottomframe.pack(side=BOTTOM, fill=X)
   # CENTER
   photo = PhotoImage(file=r"icon.png")
   photo = photo.subsample(4, 4)
   bluebutton = Button(bottomframe, command=cls, image=photo, bg="black")
   bluebutton.pack(side=BOTTOM, expand=True, fill=X)
   # bluebutton.grid(column=0,row=0,columnspan=3)

   root1.mainloop()

def main_():

   def no_user():
      root.destroy()
      main2()
   root = Tk()
   root.title("HOME")
   root.maxsize(350,600)
   root.minsize(350,600)
   root.configure(bg='white')
   frame=Frame(root,highlightbackground="black",highlightthickness=25)
   frame.pack(side=TOP,expand = True, fill = BOTH)

   Fullname=StringVar()
   Age=IntVar()
   Contact = IntVar()
   Salary=IntVar()
   Location= StringVar()

   def database():
      #conn = sqlite3.connect('Form.db')
      #with conn:
         #cursor=conn.cursor()
      cursor.execute('CREATE TABLE IF NOT EXISTS User (Fullname TEXT,Age TEXT,Contact TEXT,Salary TEXT,Location TEXT)')
    #########  cursor.execute('INSERT INTO Employee (FullName,Age,Contact,Salary,Location) VALUES(?,?,?,?,?)',(name1,age,con,sal,loc,))
      #conn.commit()
      label_1 = Label(root, text="NAME",width=5,font=("bold", 10))
      label_1.place(x=80,y=130)

      def SpeakFullname():
         r = sr.Recognizer()
         with sr.Microphone() as source:
            print("Speak:")
            audio = r.listen(source)
            try:
               text = r.recognize_google(audio)
               Fullname.set(text)
            except:
               messagebox.showerror("Try Again","Sorry could not recognize what you said")

      def SpeakAge():
         r = sr.Recognizer()
         with sr.Microphone() as source:
            print("Speak:")
            audio = r.listen(source)
            try:
               text = r.recognize_google(audio)
               Age.set(text)
            except:
               messagebox.showerror("Try Again","Sorry could not recognize what you said")

      def SpeakContact():
         r = sr.Recognizer()
         with sr.Microphone() as source:
            print("Speak:")
            audio = r.listen(source)
            try:
               text = r.recognize_google(audio)
               Contact.set(text)
            except:
               messagebox.showerror("Try Again","Sorry could not recognize what you said")

      def SpeakSalary():
         r = sr.Recognizer()
         with sr.Microphone() as source:
            print("Speak:")
            audio = r.listen(source)
            try:
               text = r.recognize_google(audio)
               Salary.set(text)
            except:
               messagebox.showerror("Try Again","Sorry could not recognize what you said")

      entry_1 = Entry(root, textvariable=Fullname)
      entry_1.place(x=180, y=130)
      Button(root, text='*', command=SpeakFullname).place(x=300, y=130)

      label_2 = Label(root, text="AGE",width=6,font=("bold", 10))
      label_2.place(x=68,y=180)

      entry_2 = Entry(root, textvariable=Age)
      entry_2.place(x=180, y=180)
      Button(root, text='*', command=SpeakAge).place(x=300, y=180)

      label_3 = Label(root, text="CONTACT",width=9,font=("bold", 10))
      label_3.place(x=70,y=230)

      entry_3 = Entry(root, textvariable=Contact)
      entry_3.place(x=180, y=230)
      Button(root, text='*', command=SpeakContact).place(x=300, y=230)

      label_4 = Label(root, text="SALARY",width=8,font=("bold", 10))
      label_4.place(x=70,y=280)

      entry_4 = Entry(root, textvariable=Salary)
      entry_4.place(x=180, y=280)
      Button(root, text='*', command=SpeakSalary).place(x=300, y=280)


      label_5 = Label(root, text="LOCATION",width=10,font=("bold", 10))
      label_5.place(x=70,y=330)

      #list1 = ['Adavapakkam','Angambakkam','Ariyaperumpakkam','Asoor','Chinnalambadi','Damal','Illalur','Kalpattu','Karalappakkam','Kilakkadi','Kottavakkam','Kottur','Melmaduramangalam','Nanjeepuram','Narapakkam','Nathanallur','Pondur','Putheri','Sethupattu','Sirukalathur','Thammanur','Tharapakkam','Thenialam','Thenneri','Thodur','Vaipoor','Vellacheri','Vilangadupakkam'];
      droplist= Combobox(root, text="choose location", height=5, width=15,
                        values= ['Adavapakkam','Angambakkam','Ariyaperumpakkam','Asoor','Chinnalambadi','Damal','Illalur','Kalpattu','Karalappakkam','Kilakkadi','Kottavakkam','Kottur','Melmaduramangalam','Nanjeepuram','Narapakkam','Nathanallur','Pondur','Putheri','Sethupattu','Sirukalathur','Thammanur','Tharapakkam','Thenialam','Thenneri','Thodur','Vaipoor','Vellacheri','Vilangadupakkam'])
      droplist.place(x=180,y=330)

      #add your code here
      #delete the 'pass' and then start writing
      #inserting data and displaying in the new window
      def func():
         a = entry_1.get()
         b = entry_2.get()
         c = entry_3.get()
         f = entry_4.get()
         g = droplist.get()
         x = (a, b, c, f, g)
         d = {'col1': [b], 'col2': [f]}
         df = pd.DataFrame(data=d)
         result = clf.predict(df)
         result = np.array_str(result)
         result = result[2:-2]
         z = str(result)
         ans=messagebox.askyesno("Confirm","Predicted job for "+a+" is "+z+" \n Do you want to confirm?")
         if ans:#enetering into the table
            def otp():

               rooto = Tk()
               rooto.title("OTP AUTHENTICATION")
               n = random.randint(1000, 9999)
               print(n)
               rooto.maxsize(250, 150)
               rooto.minsize(250, 150)
               OTP = IntVar()

               def fun():
                  client = Client("AC68b58731071af626a9f982d4e0f0065d", "161186c5469ab04a3526e23c68016bef")
                  mess = client.messages.create(body="Your OTP is " + str(n), from_="+16122948683", to="+918420315480")
                  print(mess.sid)

               label_1 = Label(rooto, text="Click to get an OTP", width=14, font=("bold", 10))
               label_1.place(x=75, y=10)
               Button(rooto, text='SEND', width=10, bg='black', fg='white', command=fun).place(x=95, y=30)

               label_2 = Label(rooto, text="Enter the OTP", width=12, font=("bold", 10))
               label_2.place(x=20, y=80)

               entry_2 = Entry(rooto, width=10)
               entry_2.place(x=125, y=80)

               def fun1():
                  xx = entry_2.get()
                  print(xx)
                  n1 = str(n)
                  if n1 == xx:
                     rooto.destroy()
                     messagebox.showinfo("Successful", "OTP Matched successfully")

                     cursor.execute("insert into User values(?,?,?,?,?)", x)
                     messagebox.showinfo("Successful", "Data entered successfully")
                     # conn.commit()
                     RES = messagebox.askyesno("Insert More", "Do you want to insert more users?", )
                     if RES:
                        database()
                     else:
                        cursor.execute("select Fullname, Age, Contact, Salary, Location from User")
                        messagebox.showinfo("Required Details", cursor.fetchall())
                        root.destroy()
                        main2()
                  else:
                     messagebox.showerror("Try Again", "Wrong OTP,Please Try Again")
                     rooto.destroy()
                     database()

               label_3 = Label(rooto, text="Click to verify", width=14, font=("bold", 10))
               label_3.place(x=15, y=120)
               Button(rooto, text='VERIFY', width=10, bg='black', fg='white', command=fun1).place(x=120, y=120)

               rooto.mainloop()
            otp()
            #conn = sqlite3.connect('Form.db')



            #age_=e[0][1]
            #salary_=e[0][3]
            #d1 = {'col1': [age_], 'col2': [salary_]}
            #df1 = pd.DataFrame(data=d1)
            #result1 = clf.predict(df1)
            #txt="Predicted job for " + e[0][0] + " is " + result1
            #label_01 = Label(root, text=txt, width=30, font=("Calibri 10 bold", 10))
            #label_01.place(x=30, y=420)
            #tkinter.messagebox.showinfo("Job Prediction", cursor.fetchall())

      def track():
         def cls1():
            troot.destroy()
            main_()

         def tr():
            a = entry_1.get()
            b = entry_2.get()
            cursor.execute("select rowid,Fullname from User where Fullname=(?) and Contact=(?)",(a,b,))
            e=cursor.fetchall()
            token=str(e[0][0])
            tname=e[0][1]
            messagebox.showinfo("Details", tname +", You are number " + token + " in the queue.")
         root.destroy()
         troot=Tk()
         troot.maxsize(350, 600)
         troot.minsize(350, 600)
         troot.configure(bg='white')
         tframe = Frame(troot, highlightbackground="black", highlightthickness=25)
         tframe.pack(side=TOP, expand=True, fill=BOTH)

         label_1 = Label(troot, text="NAME", width=5, font=("bold", 10))
         label_1.place(x=80, y=130)

         entry_1 = Entry(troot)
         entry_1.place(x=180, y=130)

         label_2 = Label(troot, text="CONTACT", width=7, font=("bold", 10))
         label_2.place(x=68, y=180)

         entry_2 = Entry(troot)
         entry_2.place(x=180, y=180)

         Button(troot, text='TRACK', command=tr, width=10, bg='black', fg='white').place(x=100, y=230)

         tbottomframe = Frame(tframe, highlightbackground="black", highlightthickness=5)
         tbottomframe.pack(side=BOTTOM, fill=X)
         # CENTER
         tphoto = PhotoImage(file=r"icon.png")
         tphoto = tphoto.subsample(4, 4)
         bluebutton = Button(tbottomframe,command=cls1, image=tphoto, bg="black")
         bluebutton.pack(side=BOTTOM, expand=True, fill=X)
         # bluebutton.grid(column=0,row=0,columnspan=3)

         troot.mainloop()
      def helpp():
         rooth = Tk()
         rooth.geometry('550x250')
         rooth.resizable(0, 0)
         rooth.config(bg='ghost white')
         rooth.title('INTELLIGENT JOB NAVIGATION SYSTEM FOR RURAL DEVELOPMENT')
         Label(rooth, text='Dear Candidate, please type out your queries / personal details below.', font='arial 9',
               bg='white smoke').pack()
         Label(rooth, text='அன்புள்ள வேட்பாளர், உங்கள் கேள்விகளை / தனிப்பட்ட விவரங்களை கீழே எழுதுங்கள்.', font='arial 8',
               bg='white smoke').pack()
         Label(rooth, text='எந்த உதவிக்கும், மையத்தில் உள்ள நிர்வாகியைத் தொடர்பு கொள்ளுங்கள்.', font='arial 9',
               bg='white smoke').pack(side=BOTTOM)
         Label(rooth, text='For any further assistance, contact the admin at the center.', font='arial 8',
               bg='white smoke').pack(side=BOTTOM)
         Label(rooth, text='Enter Text - உங்கள் கேள்விகள் அல்லது தனிப்பட்ட விவரங்களை எழுதுங்கள் ', font='arial 9 bold',
               bg='white smoke').place(x=20, y=60)
         Msg1 = StringVar()
         entry_field = Entry(rooth, textvariable=Msg1, width='85')
         entry_field.place(x=20, y=100)

         def Text_to_speech():
            language = "ta"
            myobj = gTTS(text=entry_field.get(),
                         lang=language,
                         slow=False)
            myobj.save("SpeechAudio.wav")
            os.system("SpeechAudio.wav")

         def Exit():
            rooth.destroy()

         def Reset():
            a=entry_field.get()
            file = openpyxl.load_workbook("Queries.xlsx")
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=a)
            file.save("Queries.xlsx")

         Button(rooth, text="PLAY", font='arial 15 bold', command=Text_to_speech, bg='green', width=4).place(x=25, y=140)
         Button(rooth, text='EXIT', font='arial 15 bold', command=Exit, bg='OrangeRed1').place(x=100, y=140)
         Button(rooth, text='SUBMIT', font='arial 15 bold', command=Reset, bg='yellow').place(x=175, y=140)
         rooth.mainloop()
      Button(root, text='HELP', command=helpp, width=5, bg='red', fg='black').place(x=155, y=80)
      Button(root, text='REGISTER', command=func, width=8, bg='black', fg='white').place(x=230, y=380)
      Button(root, text='START', command=no_user, width=6, bg='black', fg='white').place(x=140, y=380)
      Button(root, text='TRACK', command=track, width=6, bg='black', fg='white').place(x=50, y=380)
   database()

   bottomframe = Frame(frame,highlightbackground="black",highlightthickness=5)
   bottomframe.pack( side = BOTTOM ,fill=X)
   #CENTER
   photo = PhotoImage(file = r"icon.png")
   photo=photo.subsample(4, 4)
   bluebutton = Button(bottomframe, image = photo,bg="black")
   bluebutton.pack( side = BOTTOM ,expand = True, fill = X)
   #bluebutton.grid(column=0,row=0,columnspan=3)

   root.mainloop()



main_()