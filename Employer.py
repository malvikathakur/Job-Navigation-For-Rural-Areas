from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
from openpyxl import Workbook
import joblib
clf = joblib.load('model.pickle')
from tkinter.ttk import Combobox
import sqlite3
file=pathlib.Path("Employer.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Name"
    sheet["B1"] = "Contact"
    sheet["C1"] = "Location"
    sheet["D1"] = "Job Type"
    sheet["E1"] = "Age Required"
    sheet["F1"] = "Salary"
    file.save("Employer.xlsx")
root = Tk()
root.maxsize(350, 600)
root.minsize(350, 600)
root.configure(bg='white')
root.title('JOB PROVIDER DETAILS')
frame = Frame(root, highlightbackground="black", highlightthickness=25)
frame.pack(side=TOP, expand=True, fill=BOTH)
Fullname = StringVar()
Contact = IntVar()
Age = IntVar()
JOB = StringVar()
Salary = IntVar()
Location = StringVar()
conn = sqlite3.connect('Form.db')
cursor = conn.cursor()
def database():
    with conn:
        cursor=conn.cursor()
        cursor.execute( 'CREATE TABLE IF NOT EXISTS Employer (Fullname TEXT,Contact TEXT, Age TEXT,JOB TEXT,Salary TEXT,Location TEXT)')
       # cursor.execute('INSERT INTO Employer (FullName,Contact,Age,JOB,Salary,Location) VALUES(?,?,?,?,?,?)',(name1,con,age,job,sal,loc,))
        conn.commit()
    label_0 = Label(root, text="EMPLOYER DETAILS", width=28, font=("bold", 12))
    label_0.place(x=60, y=100)

    label_1 = Label(root, text="NAME", width=5, font=("bold", 10))
    label_1.place(x=80, y=130)

    entry_1 = Entry(root)
    entry_1.place(x=180, y=130)

    label_2 = Label(root, text="CONTACT", width=7, font=("bold", 10))
    label_2.place(x=80, y=165)

    entry_2 = Entry(root)
    entry_2.place(x=180, y=160)
    label_C = Label(root,text="JOB TYPE", width=12, font=("bold", 10))
    label_C.place(x=60, y=190)
    droplist1 = Combobox(root, text="JOB TYPE", height=5, width=15,
                        values=['Agriculture','Construction','Painter','Babysitter','House-Maid','Sanitation','Cooking','Security','Laundry','Electrician','Food-Delivery','Sweeper','Housekeeping'])
    droplist1.place(x=180, y=190)

    label_4 = Label(root, text="REQUIRED AGE", width=14, font=("bold", 10))
    label_4.place(x=70, y=230)

    entry_4 = Entry(root)
    entry_4.place(x=180, y=230)

    label_5 = Label(root, text="SALARY", width=8, font=("bold", 10))
    label_5.place(x=70, y=260)

    entry_5 = Entry(root)
    entry_5.place(x=180, y=260)

    label_6 = Label(root, text="LOCATION", width=10, font=("bold", 10))
    label_6.place(x=70, y=290)
    droplist = Combobox(root, text="choose location", height=5, width=15,
                        values=['Adavapakkam', 'Angambakkam', 'Ariyaperumpakkam', 'Asoor', 'Chinnalambadi', 'Damal','Illalur', 'Kalpattu', 'Karalappakkam', 'Kilakkadi', 'Kottavakkam', 'Kottur','Melmaduramangalam', 'Nanjeepuram', 'Narapakkam', 'Nathanallur', 'Pondur', 'Putheri','Sethupattu', 'Sirukalathur', 'Thammanur', 'Tharapakkam', 'Thenialam', 'Thenneri','Thodur', 'Vaipoor', 'Vellacheri', 'Vilangadupakkam'])
    droplist.place(x=180, y=290)

    def func():
        a = entry_1.get()
        b = entry_2.get()
        c = droplist1.get()
        f = entry_4.get()
        h = entry_5.get()
        g = droplist.get()
        z = (a, b, c, f, h, g)
        file=openpyxl.load_workbook("Employer.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=a)
        sheet.cell(column=2, row=sheet.max_row, value=b)
        sheet.cell(column=3, row=sheet.max_row, value=g)
        sheet.cell(column=4, row=sheet.max_row, value=c)
        sheet.cell(column=5, row=sheet.max_row, value=f)
        sheet.cell(column=6, row=sheet.max_row, value=h)
        file.save("Employer.xlsx")
        cursor.execute("insert into Employer values(?,?,?,?,?,?)", (z),)
        #conn.commit()
        RES=messagebox.askyesno("Insert More", "Do you want to insert more employers?",)
        if RES:
            database()
        else:
            cursor.execute("select * from Employer ")
            messagebox.showinfo("Required Details", cursor.fetchall())
            root.destroy()

    Button(root, text='ENTER',command=func, width=10, bg='blue', fg='white').place(x=180, y=320)
database()

bottomframe = Frame(frame, highlightbackground="black", highlightthickness=5)
bottomframe.pack(side=BOTTOM, fill=X)
photo = PhotoImage(file=r"icon.png")
photo = photo.subsample(4, 4)
bluebutton = Button(bottomframe, image=photo, bg="black")
bluebutton.pack(side=BOTTOM, expand=True, fill=X)
root.mainloop()
